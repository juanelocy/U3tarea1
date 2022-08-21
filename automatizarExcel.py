'''Librerias para el uso del problema'''
from openpyxl import Workbook
from openpyxl.styles import Font
import time 

'''Instanciacion del libro o docuemento en excel'''
book=Workbook()
'''Llamando al metodo active para que funcione las hojas, creado una hoja dentro del documento'''
hoja1=book.active
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja1['A1']='Id'
'''funcion .font para agregar color a una celda elegida'''
hoja1['A1'].font=Font(color='FF0000', bold=True)
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja1['B1']='Nombres'
'''funcion .font para agregar color a una celda elegida'''
hoja1['B1'].font=Font(color='FF0000', bold=True)
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja1['C1']='Apellidos'
'''funcion .font para agregar color a una celda elegida'''
hoja1['C1'].font=Font(color='FF0000', bold=True)
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja1['D1']='Numero de telefono'
'''Funcion .merge_cells para unir celdas'''
hoja1.merge_cells('D1:E1')
'''funcion .font para agregar color a una celda elegida'''
hoja1['D1'].font=Font(color='FF0000', bold=True)
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja1['F1']='Fecha'
'''Asignando a una variable el uso de la funcion time.strtime que sirve para ingresar la fecha'''
fecha=time.strftime('%x')
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja1['F2'] = fecha

'''Funcion .create_sheet  para crear otra hoja dentro del mismo documento agregando el nombre de la misma'''
hoja2=book.create_sheet('Hoja 2')
'''Se ingresa la celda que es igual al texto que se quiera ingresar'''
hoja2['A1']='Creacion de Segunda hoja'

'''Funcion .save para guardar dentro del documento elegido'''
book.save('ExcelDePrueba.xlsx')